import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

METRICS = ["PSNR", "SSIM", "LPIPS"]
DIV2K_RE = re.compile(r"^DIV2K\d{2}$")


def div2k_sort_key(name):
    if isinstance(name, str) and DIV2K_RE.match(name):
        return int(name.replace("DIV2K", ""))
    return 9999


def setting_sort_key(setting_name):
    setting_name = str(setting_name)
    optimizer_order = {
        "adam": 0,
        "muon": 1,
        "lr-sign": 2,
        "lr_sign10_rsclF": 3,
        "auto_cos_inc_rank": 4,
        "auto_cos_inc": 4,
    }
    head = setting_name.split("_")[0]
    for opt, order in optimizer_order.items():
        if setting_name.startswith(opt):
            return (order, setting_name)
    return (999, setting_name)


def find_readme_files(root_dir: Path):
    return sorted(
        [p for p in root_dir.rglob("*") if p.is_file() and p.name.lower() == "readme.txt"],
        key=lambda p: str(p),
    )


def parse_readme(readme_path: Path):
    result = {metric: None for metric in METRICS}
    try:
        text = readme_path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return result

    for metric in METRICS:
        match = re.search(
            rf"\b{metric}\b\s*[:=]\s*([+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?)",
            text,
            flags=re.IGNORECASE,
        )
        if match:
            try:
                result[metric] = float(match.group(1))
            except ValueError:
                result[metric] = None
    return result


def parse_readme_path(root_dir: Path, readme_path: Path):
    """
    Supported structures:
      1) setting/DIV2Kxx/model/optimizer/Readme.txt
      2) setting/DIV2Kxx/model/optimizer/DIV2Kyy/Readme.txt
    """
    try:
        parts = readme_path.relative_to(root_dir).parts
    except ValueError:
        return None

    if len(parts) >= 5 and DIV2K_RE.match(parts[1]) and parts[-1].lower() == "readme.txt":
        setting = parts[0]
        outer_image = parts[1]
        experiment = parts[2]
        optimizer = parts[3]

        if len(parts) >= 6 and DIV2K_RE.match(parts[4]):
            image = parts[4]
            path_type = "nested_image_readme"
        else:
            image = outer_image
            path_type = "direct_readme"

        return {
            "Setting": setting,
            "OuterImage": outer_image,
            "Experiment": experiment,
            "Optimizer": optimizer,
            "Image": image,
            "PathType": path_type,
            "ReadmePath": str(readme_path),
        }

    return None


def collect_results(root_dir: Path, setting_filter=None, experiment_filter=None, optimizer_filter=None):
    rows = []
    diagnostics = []

    readmes = find_readme_files(root_dir)
    for readme_path in readmes:
        parsed = parse_readme_path(root_dir, readme_path)
        if parsed is None:
            diagnostics.append({
                "Status": "skipped_path_not_matching_expected_structure",
                "ReadmePath": str(readme_path),
                "Reason": "expected setting/DIV2Kxx/model/optimizer[/DIV2Kyy]/Readme.txt",
            })
            continue

        if setting_filter and parsed["Setting"] not in setting_filter:
            diagnostics.append({"Status": "skipped_setting_filter", "ReadmePath": str(readme_path), "Reason": parsed["Setting"]})
            continue
        if experiment_filter and parsed["Experiment"] not in experiment_filter:
            diagnostics.append({"Status": "skipped_experiment_filter", "ReadmePath": str(readme_path), "Reason": parsed["Experiment"]})
            continue
        if optimizer_filter and parsed["Optimizer"] not in optimizer_filter:
            diagnostics.append({"Status": "skipped_optimizer_filter", "ReadmePath": str(readme_path), "Reason": parsed["Optimizer"]})
            continue

        metrics = parse_readme(readme_path)
        if all(metrics[m] is None for m in METRICS):
            diagnostics.append({
                "Status": "skipped_metric_not_found",
                "ReadmePath": str(readme_path),
                "Reason": "PSNR/SSIM/LPIPS not found",
            })
            continue

        rows.append({**parsed, **metrics})
        diagnostics.append({"Status": "included", "ReadmePath": str(readme_path), "Reason": "ok"})

    return rows, diagnostics


def sort_raw_df(raw_df):
    if raw_df.empty:
        return raw_df
    raw_df = raw_df.copy()
    raw_df["SettingSort"] = raw_df["Setting"].map(setting_sort_key)
    raw_df["OuterSort"] = raw_df["OuterImage"].map(div2k_sort_key)
    raw_df["ImageSort"] = raw_df["Image"].map(div2k_sort_key)
    raw_df = raw_df.sort_values(["SettingSort", "Experiment", "Optimizer", "OuterSort", "ImageSort", "ReadmePath"])
    raw_df = raw_df.drop(columns=["SettingSort", "OuterSort", "ImageSort"])
    return raw_df.reset_index(drop=True)


def make_summary(raw_df):
    columns = [
        "Setting", "Experiment", "Optimizer", "Count", "Image_Count",
        "PSNR_mean", "PSNR_std", "SSIM_mean", "SSIM_std", "LPIPS_mean", "LPIPS_std",
    ]
    if raw_df.empty:
        return pd.DataFrame(columns=columns)

    summary = (
        raw_df.groupby(["Setting", "Experiment", "Optimizer"], dropna=False)
        .agg(
            Count=("Image", "count"),
            Image_Count=("Image", "nunique"),
            PSNR_mean=("PSNR", "mean"),
            PSNR_std=("PSNR", "std"),
            SSIM_mean=("SSIM", "mean"),
            SSIM_std=("SSIM", "std"),
            LPIPS_mean=("LPIPS", "mean"),
            LPIPS_std=("LPIPS", "std"),
        )
        .reset_index()
    )
    for col in columns:
        if col.endswith("_mean") or col.endswith("_std"):
            summary[col] = pd.to_numeric(summary[col], errors="coerce").round(6)
    summary["SettingSort"] = summary["Setting"].map(setting_sort_key)
    summary = summary.sort_values(["SettingSort", "Experiment", "Optimizer"]).drop(columns=["SettingSort"])
    return summary[columns]


def make_by_image_wide(raw_df, metric, image_names):
    columns = ["Setting", "Experiment", "Optimizer", *image_names, "Average"]
    if raw_df.empty:
        return pd.DataFrame(columns=columns)

    pivot = raw_df.pivot_table(
        index=["Setting", "Experiment", "Optimizer"],
        columns="Image",
        values=metric,
        aggfunc="mean",
    )
    pivot = pivot.reindex(columns=image_names)
    pivot["Average"] = pivot.mean(axis=1)
    pivot = pivot.reset_index()
    pivot["SettingSort"] = pivot["Setting"].map(setting_sort_key)
    pivot = pivot.sort_values(["SettingSort", "Experiment", "Optimizer"]).drop(columns=["SettingSort"])

    for col in image_names + ["Average"]:
        pivot[col] = pd.to_numeric(pivot[col], errors="coerce").round(6)
    return pivot[columns]


def safe_sheet_name(name):
    for ch in ["\\", "/", "*", "?", ":", "[", "]"]:
        name = name.replace(ch, "_")
    return name[:31]


def unique_sheet_name(base_name, used_names):
    base = safe_sheet_name(base_name)
    if base not in used_names:
        used_names.add(base)
        return base
    idx = 2
    while True:
        suffix = f"_{idx}"
        candidate = safe_sheet_name(base[:31 - len(suffix)] + suffix)
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        idx += 1


def autosize_worksheet(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def style_worksheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    autosize_worksheet(ws)


def parse_csv_list(text):
    if text is None or text == "":
        return None
    return {x.strip() for x in text.split(",") if x.strip()}


def main():
    parser = argparse.ArgumentParser(description="Summarize SISR grid Readme.txt metrics into an Excel file.")
    parser.add_argument("--root_dir", type=Path, default=Path("logs/sisr_grid"))
    parser.add_argument("--output_xlsx", type=Path, default=None)
    parser.add_argument("--settings", type=str, default=None, help="Comma-separated setting filter.")
    parser.add_argument("--experiments", type=str, default=None, help="Comma-separated model/experiment filter.")
    parser.add_argument("--optimizers", type=str, default=None, help="Comma-separated optimizer folder filter.")
    parser.add_argument("--image_start", type=int, default=1)
    parser.add_argument("--image_end", type=int, default=30)
    args = parser.parse_args()

    root_dir = args.root_dir
    output_xlsx = args.output_xlsx or (root_dir / "metrics_summary.xlsx")
    image_names = [f"DIV2K{i:02d}" for i in range(args.image_start, args.image_end + 1)]

    if not root_dir.exists():
        print(f"[오류] ROOT_DIR가 존재하지 않습니다: {root_dir}")
        return

    rows, diagnostics = collect_results(
        root_dir,
        setting_filter=parse_csv_list(args.settings),
        experiment_filter=parse_csv_list(args.experiments),
        optimizer_filter=parse_csv_list(args.optimizers),
    )

    columns = ["Setting", "OuterImage", "Experiment", "Optimizer", "Image", "PSNR", "SSIM", "LPIPS", "PathType", "ReadmePath"]
    raw_df = pd.DataFrame(rows, columns=columns)
    raw_df = sort_raw_df(raw_df)

    if not raw_df.empty:
        raw_df = raw_df.drop_duplicates(
            subset=["Setting", "OuterImage", "Experiment", "Optimizer", "Image"],
            keep="last",
        ).reset_index(drop=True)
        for metric in METRICS:
            raw_df[metric] = pd.to_numeric(raw_df[metric], errors="coerce").round(6)

    diag_df = pd.DataFrame(diagnostics, columns=["Status", "ReadmePath", "Reason"])
    summary_df = make_summary(raw_df)
    by_image = {metric: make_by_image_wide(raw_df, metric, image_names) for metric in METRICS}

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    used_sheet_names = set()
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        raw_df.to_excel(writer, sheet_name=unique_sheet_name("all_results", used_sheet_names), index=False)
        summary_df.to_excel(writer, sheet_name=unique_sheet_name("summary", used_sheet_names), index=False)
        for metric, df in by_image.items():
            df.to_excel(writer, sheet_name=unique_sheet_name(f"{metric}_by_image", used_sheet_names), index=False)
        diag_df.to_excel(writer, sheet_name=unique_sheet_name("diagnostics", used_sheet_names), index=False)

        for ws in writer.book.worksheets:
            style_worksheet(ws)

    print(f"[완료] 엑셀 파일 저장: {output_xlsx}")
    print(f"[정보] 발견된 Readme.txt 수: {len(find_readme_files(root_dir))}")
    print(f"[정보] 엑셀에 포함된 row 수: {len(raw_df)}")
    if raw_df.empty:
        print("[경고] 포함된 row가 0개입니다. diagnostics 시트를 확인하세요.")
    else:
        print("[정보] 포함된 setting 수:", raw_df["Setting"].nunique())
        print("[정보] 포함된 experiment:", sorted(raw_df["Experiment"].unique()))
        print("[정보] 포함된 optimizer:", sorted(raw_df["Optimizer"].unique()))


if __name__ == "__main__":
    main()
